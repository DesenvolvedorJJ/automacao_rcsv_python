import pandas as pd
import glob
from openpyxl import Workbook
from openpyxl.styles import Font

# Obtém todos os arquivos CSV no diretório especificado
csv_files = glob.glob('mes/*.csv')

# Lista para armazenar os totais de produção de cada mês
totais_producao = []

# Processa cada arquivo CSV
for file in csv_files:
    try:
        df = pd.read_csv(file, on_bad_lines='skip')
        total_producao = df['Producao'].sum()
        totais_producao.append(total_producao)
    except pd.errors.ParserError as e:
        print(f"Erro ao ler o CSV {file}: {e}")
        totais_producao.append(0)  # Adiciona 0 em caso de erro

# Cria um novo Workbook
wb = Workbook()
ws = wb.active

# Adiciona o cabeçalho
ws['A1'] = 'Mês'
ws['B1'] = 'Quantidade total do mês'

# Formata a célula A1 e B1 como negrito
bold_font = Font(bold=True)
ws['A1'].font = bold_font
ws['B1'].font = bold_font

# Adiciona os totais de produção de cada mês na planilha
meses = ['Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 'Julho']
for i, total in enumerate(totais_producao, start=1):
    ws[f'A{i+1}'] = meses[i-1]
    ws[f'B{i+1}'] = total

# Salva o arquivo Excel
wb.save('soma_total.xlsx')
