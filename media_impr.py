import pandas as pd
import glob
from openpyxl import Workbook
from openpyxl.styles import Font

# Obtém todos os arquivos CSV no diretório especificado
csv_files = glob.glob('mes/*.csv')

# Processa cada arquivo CSV
for file in csv_files:
    # Obtém o nome do mês a partir do nome do arquivo
    mes = file.split('/')[1].split('.')[0]
    
    try:
        df = pd.read_csv(file, on_bad_lines='skip')
        
        # Verifica se a coluna 'Producao' e 'Site' estão presentes
        if 'Producao' in df.columns and 'Site' in df.columns:
            # Agrupa e soma os valores da coluna 'Producao' por 'Site'
            grouped_df = df.groupby('Site')['Producao'].sum().reset_index()
        else:
            print(f"Colunas 'Producao' e/ou 'Site' não encontradas em {file}")
            grouped_df = pd.DataFrame(columns=['Site', 'Producao'])
        
        # Cria um novo Workbook
        wb = Workbook()
        ws = wb.active

        # Adiciona o cabeçalho
        ws['A1'] = 'Site'
        ws['B1'] = 'Quantidade total'

        # Formata a célula A1 e B1 como negrito
        bold_font = Font(bold=True)
        ws['A1'].font = bold_font
        ws['B1'].font = bold_font

        # Adiciona os dados do DataFrame na planilha
        for index, row in grouped_df.iterrows():
            ws[f'A{index+2}'] = row['Site']
            ws[f'B{index+2}'] = row['Producao']
        
        # Salva o arquivo Excel com o nome do mês
        wb.save(f'{mes}.xlsx')
    
    except pd.errors.ParserError as e:
        print(f"Erro ao ler o CSV {file}: {e}")
