import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
import glob
import os

# Obtém todos os arquivos CSV no diretório especificado
csv_files = glob.glob('mes/*.csv')

# Cria o diretório de saída se não existir
output_dir = 'Média mensal por fabricante'
os.makedirs(output_dir, exist_ok=True)

# Cria o diretório de saída se não existir
output_dir = 'Média mensal por fabricante'
os.makedirs(output_dir, exist_ok=True)

# Processa cada arquivo CSV
for file in csv_files:
    # Obtém o nome do mês a partir do nome do arquivo
    mes = os.path.basename(file).split('.')[0]
    
    try:
        df = pd.read_csv(file, on_bad_lines='skip')
        
        # Verifica se as colunas necessárias estão presentes
        if 'Producao' in df.columns and 'Fabricante' in df.columns and 'Endereço de IP' in df.columns:
            # Trata IP vazio como 'DESCONHECIDO' para contagem
            df['Endereço de IP'] = df['Endereço de IP'].fillna('DESCONHECIDO')
            
            # Agrupa os dados por fabricante e calcula a soma da produção e a quantidade de impressoras
            grouped_df = df.groupby('Fabricante').agg(
                Qtde_Impressao=('Producao', 'sum'),
                Qtde_Impressoras=('Endereço de IP', 'nunique')
            ).reset_index()
            
            # Calcula a média de produção por fabricante
            grouped_df['Média'] = grouped_df['Qtde_Impressao'] / grouped_df['Qtde_Impressoras']
            
            # Calcula totais gerais
            total_general = {
                'Fabricante': 'TOTAL',
                'Qtde_Impressao': grouped_df['Qtde_Impressao'].sum(),
                'Qtde_Impressoras': grouped_df['Qtde_Impressoras'].sum(),
                'Média': grouped_df['Qtde_Impressao'].sum() / grouped_df['Qtde_Impressoras'].sum() if grouped_df['Qtde_Impressoras'].sum() > 0 else 0
            }
            
            # Adiciona a linha de total ao DataFrame
            total_df = pd.DataFrame([total_general])
            grouped_df = pd.concat([grouped_df, total_df], ignore_index=True)
            
            # Cria um novo Workbook
            wb = Workbook()
            ws = wb.active

            # Adiciona o cabeçalho
            ws['A1'] = 'Fabricante'
            ws['B1'] = 'Impressões'
            ws['C1'] = 'Qtde.Impressoras'
            ws['D1'] = 'Média'

            # Formata a célula A1 até D1 como negrito
            bold_font = Font(bold=True)
            for cell in ws['1:1']:
                cell.font = bold_font

            # Adiciona os dados do DataFrame na planilha
            for index, row in grouped_df.iterrows():
                ws[f'A{index+2}'] = row['Fabricante']
                ws[f'B{index+2}'] = row['Qtde_Impressao']
                ws[f'C{index+2}'] = row['Qtde_Impressoras']
                ws[f'D{index+2}'] = row['Média']
            
            # Salva o arquivo Excel com o nome do mês
            wb.save(os.path.join(output_dir, f'{mes}_media_producao_por_fabricante.xlsx'))
        
        else:
            print(f"Colunas necessárias não encontradas em {file}")

    except pd.errors.ParserError as e:
        print(f"Erro ao ler o CSV {file}: {e}")
