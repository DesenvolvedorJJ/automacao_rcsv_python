import pandas as pd
import openpyxl
from openpyxl.chart import BarChart, Reference
import os

def create_excel_with_chart(dataframe, output_path):
    # Remover a linha "Total" do DataFrame, se existir
    dataframe = dataframe[dataframe['Fabricante'] != 'Total']

    # Create a new workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Write the dataframe to the Excel sheet
    for r in dataframe_to_rows(dataframe, index=False, header=True):
        sheet.append(r)

    # Create a bar chart
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "Média por Fabricante"
    chart.y_axis.title = ' '
    chart.x_axis.title = ' '
    
    # Add data to the chart
    data = Reference(sheet, min_col=4, min_row=1, max_row=sheet.max_row, max_col=4)
    categories = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.shape = 4
    chart.y_axis.majorGridlines = None
    
    # Customize the chart appearance
    #chart.style = 2
    #chart.height = 15  # Chart height
    #chart.width = 30   # Chart width

    # Remover a legenda do gráfico
    chart.legend = None
    
    # Add the chart to the worksheet
    sheet.add_chart(chart, "G2")

    # Save the workbook
    workbook.save(output_path)

def dataframe_to_rows(df, index=True, header=True):
    rows = []
    if header:
        rows.append(list(df.columns))
    for row in df.itertuples(index=index, name=None):
        rows.append(list(row))
    return rows

def process_excel_files(input_directory, output_directory):
    if not os.path.exists(output_directory):
        os.makedirs(output_directory)

    for filename in os.listdir(input_directory):
        if filename.endswith('.xlsx'):
            input_path = os.path.join(input_directory, filename)
            output_path = os.path.join(output_directory, f'{os.path.splitext(filename)[0]}_gráfico_media_por_fabricante.xlsx')

            df = pd.read_excel(input_path)
            create_excel_with_chart(df, output_path)

input_directory = 'Média mensal por fabricante'
output_directory = 'Gráficos média por fabricante'

process_excel_files(input_directory, output_directory)
